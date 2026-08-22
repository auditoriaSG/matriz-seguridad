import json
import re
import shutil
import unicodedata
import uuid
from datetime import date, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "local-data" / "new-plantilla.xlsx"
SNAPSHOT = ROOT / "local-data" / "admin-snapshot.json"
BACKUPS = ROOT / "local-data" / "backups"


def clean(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("�", "Ñ")).strip()


def key(value):
    text = unicodedata.normalize("NFKD", clean(value)).encode("ascii", "ignore").decode().upper()
    return re.sub(r"[^A-Z0-9]+", " ", text).strip()


def iso_date(value):
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = clean(value)
    match = re.match(r"(\d{4})-(\d{2})-(\d{2})", text)
    return match.group(0) if match else None


def real_person(name):
    marker = key(name)
    return bool(marker) and not any(x in marker for x in ("VACANTE", "NO DISPO", "NO NECESARIO", "PENDIENTE", "N A"))


def identity_name(name):
    return key(re.sub(r"\s*\((?:CAP|MAT|REINGRESO)\)\s*$", "", clean(name), flags=re.I))


def split_name(full_name):
    words = clean(full_name).title().split()
    if len(words) <= 2:
        return (words[0] if words else "", words[1] if len(words) > 1 else None, None)
    return (" ".join(words[:-2]), words[-2], words[-1])


def catalog_match(items, value, aliases=None):
    wanted = key(value)
    if aliases and wanted in aliases:
        wanted = key(aliases[wanted])
    for item in items:
        if key(item.get("nombre")) == wanted or key(item.get("codigo")) == wanted.replace(" ", "_"):
            return item
    for item in items:
        if wanted and (wanted in key(item.get("nombre")) or key(item.get("nombre")) in wanted):
            return item
    return None


def rows_from_block(ws, start, end, columns, sucursal, departamento=None):
    number_col, name_col, email_col, job_col, profession_col, date_col = columns
    output = []
    current_department = departamento
    for row in range(start, end + 1):
        number = ws.cell(row, number_col).value
        name = ws.cell(row, name_col).value
        job = ws.cell(row, job_col).value
        if isinstance(number, str) and not number.strip().isdigit() and not name:
            current_department = clean(number)
            continue
        if not real_person(name) or not clean(job):
            continue
        output.append({
            "nombre_completo": clean(name),
            "correo": clean(ws.cell(row, email_col).value).lower() or None,
            "puesto": clean(job),
            "profesion": clean(ws.cell(row, profession_col).value) or None,
            "fecha_ingreso": iso_date(ws.cell(row, date_col).value),
            "sucursal": clean(sucursal) or None,
            "departamento_fuente": clean(current_department) or None,
        })
    return output


def extract(ws):
    blocks = []
    # Sucursales: tres bloques horizontales por sección.
    for header_row, end_row in ((4, 23), (24, 46), (47, 61)):
        for header_col, columns in (
            (3, (2, 3, 5, 4, 6, 7)),
            (10, (9, 10, 11, 12, 13, 14)),
            (17, (16, 17, 18, 19, 20, 21)),
        ):
            header = clean(ws.cell(header_row, header_col).value)
            branch = re.sub(r"^SUCURSAL:\s*", "", header, flags=re.I)
            branch = re.split(r"\s+TERAPEUTAS?\b", branch, flags=re.I)[0].strip()
            blocks.extend(rows_from_block(ws, header_row + 2, end_row, columns, branch, "Operaciones"))

    # Corporativo y médicos.
    blocks.extend(rows_from_block(ws, 65, 110, (2, 3, 5, 4, 6, 7), "Oficina Monterrey", "Capital Humano"))
    blocks.extend(rows_from_block(ws, 64, 110, (9, 10, 11, 12, 13, 14), "Oficina Monterrey", "Médicos"))

    # Capacitación incluye la sucursal explícita en la columna V.
    for row in range(64, 110):
        name, job = ws.cell(row, 17).value, ws.cell(row, 19).value
        if real_person(name) and clean(job):
            blocks.append({
                "nombre_completo": clean(name), "correo": clean(ws.cell(row, 18).value).lower() or None,
                "puesto": clean(job), "profesion": clean(ws.cell(row, 20).value) or None,
                "fecha_ingreso": iso_date(ws.cell(row, 21).value), "sucursal": clean(ws.cell(row, 22).value) or None,
                "departamento_fuente": "Capacitación",
            })
    return blocks


def main():
    data = json.loads(SNAPSHOT.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    incoming = extract(wb["PLANTILLA "])

    timestamp = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    BACKUPS.mkdir(parents=True, exist_ok=True)
    backup = BACKUPS / f"admin-snapshot-antes-plantilla-{timestamp}.json"
    shutil.copy2(SNAPSHOT, backup)

    old = data.get("empleados", [])
    by_email = {key(x.get("correo")): x for x in old if x.get("correo")}
    by_name = {identity_name(" ".join(filter(None, [x.get("nombres"), x.get("apellido_paterno"), x.get("apellido_materno")]))): x for x in old}
    max_emp = max((int(re.search(r"\d+", x.get("numero_empleado", "0")).group()) for x in old if re.search(r"\d+", x.get("numero_empleado", ""))), default=137)

    branch_aliases = {"GOMEZ MORIN": "Gómez Morín", "MAZATLAN": "Mazatlán", "RIOJA": "Rioja"}
    job_aliases = {"ENCARGADA": "Encargada", "EJECUTIVA": "Ejecutiva", "TERAPEUTA": "Terapeuta", "TERAPEUTA MASTER": "Terapeuta Máster", "DIRECTOR MEDICO": "Director Medico", "MEDICO MTY": "Medico /Mty"}
    departments = data.get("departamentos", [])
    jobs = data.get("puestos", [])
    branches = data.get("sucursales", [])
    profiles = data.get("perfiles", [])
    result = []

    claimed_ids = set()
    for record in incoming:
        record["nombre_completo"] = re.sub(r"\s*\((?:CAP|MAT|REINGRESO)\)\s*$", "", record["nombre_completo"], flags=re.I)
        existing = by_name.get(identity_name(record["nombre_completo"]))
        if not existing:
            email_match = by_email.get(key(record.get("correo")))
            existing = email_match if email_match and email_match.get("id") not in claimed_ids else None
        if existing:
            employee = dict(existing)
            claimed_ids.add(employee.get("id"))
        else:
            max_emp += 1
            names, paternal, maternal = split_name(record["nombre_completo"])
            employee = {"id": str(uuid.uuid4()), "empleado_id": None, "numero_empleado": f"EMP-{max_emp}", "nombres": names, "apellido_paterno": paternal, "apellido_materno": maternal, "telefono": None, "estado": "ACTIVO"}
            employee["empleado_id"] = employee["id"]

        job = catalog_match(jobs, record["puesto"], job_aliases)
        job_name = job["nombre"] if job else clean(record["puesto"]).title()
        department_name = job.get("departamento") if job else record.get("departamento_fuente") or "Operaciones"
        department = catalog_match(departments, department_name)
        branch = catalog_match(branches, record.get("sucursal"), branch_aliases)

        profile_name = "Empleado"
        job_key = key(job_name)
        if "ENCARGADA" in job_key: profile_name = "Encargado"
        elif "EJECUTIVA" in job_key or "RECEPCION" in job_key: profile_name = "Recepcionista"
        elif "SISTEMA" in key(department_name): profile_name = "Soporte"
        elif any(x in key(department_name) for x in ("AUDITOR", "DIRECCION")): profile_name = "Propietario"
        elif not any(x in job_key for x in ("TERAPEUTA", "MEDICO")): profile_name = "Operaciones"
        profile = catalog_match(profiles, profile_name)

        employee.update({
            "correo": record.get("correo"), "puesto": job_name, "puesto_id": job.get("id") if job else None,
            "departamento": department.get("nombre") if department else department_name, "departamento_id": department.get("id") if department else None,
            "sucursal": branch.get("nombre") if branch else record.get("sucursal"), "sucursal_id": branch.get("id") if branch else None,
            "perfil": profile.get("nombre") if profile else profile_name, "perfil_id": profile.get("id") if profile else None,
            "profesion": record.get("profesion"), "fecha_ingreso": record.get("fecha_ingreso"),
        })
        result.append(employee)

    # Elimina duplicados del archivo fuente por nombre; distintas personas pueden compartir correo.
    unique = {}
    for employee in result:
        identity = identity_name(" ".join(filter(None, [employee.get("nombres"), employee.get("apellido_paterno"), employee.get("apellido_materno")])))
        unique[identity] = employee
    data["empleados"] = list(unique.values())
    SNAPSHOT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"empleados_fuente": len(incoming), "empleados_unicos": len(unique), "conservados": sum(1 for x in unique.values() if x.get("id") in {o.get('id') for o in old}), "nuevos": sum(1 for x in unique.values() if x.get("id") not in {o.get('id') for o in old}), "backup": str(backup)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
